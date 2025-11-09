import openpyxl
import mysql.connector
import os

# Conectar a la base de datos
db_url = os.environ.get('DATABASE_URL', '')
if db_url.startswith('mysql://'):
    db_url = db_url.replace('mysql://', '')
    
parts = db_url.split('@')
user_pass = parts[0].split(':')
host_db = parts[1].split('/')
host_port = host_db[0].split(':')

conn = mysql.connector.connect(
    host=host_port[0],
    port=int(host_port[1]) if len(host_port) > 1 else 3306,
    user=user_pass[0],
    password=user_pass[1] if len(user_pass) > 1 else '',
    database=host_db[1].split('?')[0]
)

cursor = conn.cursor()

# Abrir el archivo Excel
wb = openpyxl.load_workbook('/home/ubuntu/tienda-b2b/ejemplo.xlsx')
ws = wb.active

# Leer encabezados
headers = {}
for col_idx, cell in enumerate(ws[1], 1):
    if cell.value:
        headers[cell.value.strip()] = col_idx

print(f"Columnas encontradas: {list(headers.keys())}")
print(f"Total filas: {ws.max_row}")

# Las columnas relevantes son:
# 'Descripción del modelo' = nombre del producto
# 'Descripcion' = valor de la variante
# 'ciudad', 'interior', 'especial' = precios

updated_count = 0
not_found_count = 0

for row_idx in range(2, ws.max_row + 1):
    row = ws[row_idx]
    
    # Obtener valores usando los índices de columna correctos
    desc_modelo_idx = headers.get('Descripción del modelo')
    descripcion_idx = headers.get('Descripcion')
    ciudad_idx = headers.get('ciudad')
    interior_idx = headers.get('interior')
    especial_idx = headers.get('especial')
    
    if not all([desc_modelo_idx, descripcion_idx, ciudad_idx, interior_idx, especial_idx]):
        print("No se encontraron todas las columnas necesarias")
        break
    
    nombre_producto = row[desc_modelo_idx - 1].value
    valor_variante = row[descripcion_idx - 1].value
    precio_ciudad = row[ciudad_idx - 1].value
    precio_interior = row[interior_idx - 1].value
    precio_especial = row[especial_idx - 1].value
    
    if not nombre_producto or not valor_variante:
        continue
    
    # Buscar la variante en la base de datos
    cursor.execute("""
        SELECT pv.id, p.name
        FROM productVariants pv
        JOIN products p ON pv.productId = p.id
        WHERE p.name = %s AND pv.variantValue = %s
    """, (nombre_producto, valor_variante))
    
    result = cursor.fetchone()
    if result:
        variant_id = result[0]
        product_name = result[1]
        
        # Actualizar precios
        cursor.execute("""
            UPDATE productVariants
            SET precioCiudad = %s,
                precioInterior = %s,
                precioEspecial = %s
            WHERE id = %s
        """, (precio_ciudad, precio_interior, precio_especial, variant_id))
        
        updated_count += 1
        print(f"✓ {product_name} - {valor_variante}: ${precio_ciudad} / ${precio_interior} / ${precio_especial}")
    else:
        not_found_count += 1
        if not_found_count <= 5:  # Solo mostrar los primeros 5 no encontrados
            print(f"✗ No encontrado: {nombre_producto} - {valor_variante}")

conn.commit()
cursor.close()
conn.close()

print(f"\n{'='*60}")
print(f"Total variantes actualizadas: {updated_count}")
print(f"Total no encontradas: {not_found_count}")
print(f"{'='*60}")
