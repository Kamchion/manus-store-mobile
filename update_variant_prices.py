import openpyxl
import mysql.connector
import os
from decimal import Decimal

# Conectar a la base de datos
db_url = os.environ.get('DATABASE_URL', '')
# Parse DATABASE_URL: mysql://user:pass@host:port/dbname
if db_url.startswith('mysql://'):
    db_url = db_url.replace('mysql://', '')
    
parts = db_url.split('@')
user_pass = parts[0].split(':')
host_db = parts[1].split('/')
host_port = host_db[0].split(':')

conn = mysql.connector.connect(
    host=host_port[0],
    port=int(host_port[1]) if len(host_port) > 1 else 3306,
    user=user_pass[0],
    password=user_pass[1] if len(user_pass) > 1 else '',
    database=host_db[1].split('?')[0]
)

cursor = conn.cursor()

# Abrir el archivo Excel
wb = openpyxl.load_workbook('/home/ubuntu/tienda-b2b/ejemplo.xlsx')
ws = wb.active

# Leer encabezados
headers = {}
for col_idx, cell in enumerate(ws[1], 1):
    if cell.value:
        headers[cell.value.strip()] = col_idx

print(f"Columnas encontradas: {list(headers.keys())}")

# Procesar cada fila
updated_count = 0
for row_idx in range(2, ws.max_row + 1):
    row = ws[row_idx]
    
    # Obtener valores
    nombre = row[headers.get('Nombre', 1) - 1].value
    variante = row[headers.get('Variante', 2) - 1].value
    precio_ciudad = row[headers.get('Precio Ciudad', 3) - 1].value
    precio_interior = row[headers.get('Precio Interior', 4) - 1].value
    precio_especial = row[headers.get('Precio Especial', 5) - 1].value
    
    if not nombre or not variante:
        continue
    
    # Buscar la variante en la base de datos
    cursor.execute("""
        SELECT pv.id, p.id as productId
        FROM productVariants pv
        JOIN products p ON pv.productId = p.id
        WHERE p.name = %s AND pv.variantValue = %s
    """, (nombre, variante))
    
    result = cursor.fetchone()
    if result:
        variant_id = result[0]
        
        # Actualizar precios
        cursor.execute("""
            UPDATE productVariants
            SET precioCiudad = %s,
                precioInterior = %s,
                precioEspecial = %s
            WHERE id = %s
        """, (precio_ciudad, precio_interior, precio_especial, variant_id))
        
        updated_count += 1
        print(f"Actualizado: {nombre} - {variante} (Ciudad: ${precio_ciudad}, Interior: ${precio_interior}, Especial: ${precio_especial})")
    else:
        print(f"No encontrado: {nombre} - {variante}")

conn.commit()
cursor.close()
conn.close()

print(f"\nTotal de variantes actualizadas: {updated_count}")
